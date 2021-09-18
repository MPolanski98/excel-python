from openpyxl import load_workbook

target_filename = 'Sample.xlsx'

values = []

wb = load_workbook(filename=target_filename)

# ws = wb.active
ws = wb["Base"]

# for row in range(2,ws.max_row+1):  
#     for column in "K":
#         cell_name = "{}{}".format(column, row)
#         values.append(ws[cell_name].value)

values.append(ws['D3'].internal_value)
print(values)

# print(values)

